from openpyxl import Workbook
from openpyxl import load_workbook

wb=load_workbook('B.xlsx')
sheet=wb.active  #获取当前活动的active
print(sheet)

#获取行数
rows_numbers=sheet.max_row
print(rows_numbers)

#获取列数
columns_numbers=sheet.max_column
print(columns_numbers)

result=[]
for i in range(1,rows_numbers+1):
    temp=[]
    for j in range(columns_numbers):
        temp.append(sheet[i][j].value)
    result.append(temp)
print(result)

#定义一个空字典
result_key={}
for i in range(1,len(result)):
    name=result[i][0]
    number=0
    for j in range(columns_numbers):
        if result[i][j]=="缺勤":
            number=number+1
    result_key[name]=number
    #print(name,number)
print(result_key)
wb.close()

#然后读取表格A，A表中存放的是员工的基本工资表，和缺勤扣除的钱，只需要把缺勤的天数填进去就可以了

print(result_key['二狗'])
wb2=load_workbook("A.xlsx")

sheet2=wb2.active  #获取当前活动的active

#获取行数
rows_numbers2=sheet2.max_row
print(rows_numbers2)

#获取列数
columns_numbers2=sheet2.max_column
print(columns_numbers2)


for i in range(2,rows_numbers2+1):  #行从第二行取，第一行是标题
    name=sheet2[i][0].value
    print(name)
    print("缺勤天数是： ",result_key[name])
    sheet2.cell(row=i,column=4).value=result_key[name]  #ps:使用sheet.cell(row,column)写入的时候，列号是从1开始的
wb2.save('A.xlsx')






