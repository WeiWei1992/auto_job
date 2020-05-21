from openpyxl import Workbook
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
#发送多种类型的邮件
from email.mime.multipart import MIMEMultipart
import datetime

def send_email(in_data):
    msg_from = '1508691067@qq.com'  # 发送方邮箱
    passwd = 'fgaplzfksqsihdbe'

    # msg_from="1669199947@qq.com"
    # passwd="smvgfajywtnqcacd"
    to=in_data[-1]
    print(to)

    msg = MIMEMultipart()
    # conntent="这个是字符串"
    # # #把内容加进去
    # msg.attach(MIMEText(conntent,'plain','utf-8'))

    # 构造HTML
    content = '''
                    <html>
                    <body>
                    <table border='1'>
                        <tr>
                            <th>姓名</th>
                            <th>工资</th>
                            <th>Email</th>
                        </tr>
                        <tr>
                            <td>{name}</td>
                            <td>{money}</td>
                            <td>{email}</td>
                        </tr>
                        
                    </table>
                    <body>
                    <html>
                    '''.format(name=in_data[0],money=in_data[1],email=in_data[2])


    msg.attach(MIMEText(content, 'html', 'utf-8'))

    # 设置邮件主题
    msg['Subject'] = "这个是邮件主题"

    # 发送方信息
    msg['From'] = msg_from

    # 通过SSL方式发送，服务器地址和端口
    s = smtplib.SMTP_SSL("smtp.qq.com", 465)
    # 登录邮箱
    s.login(msg_from, passwd)
    # 开始发送
    s.sendmail(msg_from, to, msg.as_string())

    print("邮件发送成功")
    s.quit()



wb=load_workbook('1.xlsx')
sheet=wb.active
print(sheet)

#获取行数
rows_numbers=sheet.max_row
print(rows_numbers)

#获取列数
columns_numbers=sheet.max_column
print(columns_numbers)

result=[]
for i in range(1,rows_numbers+1):
    temp=[]
    for j in range(columns_numbers):
        temp.append(sheet[i][j].value)
    result.append(temp)
print(result)

if __name__=="__main__":
    for i in range(1,len(result)):
        send_email(result[i])


#print(sheet[0][1].value)